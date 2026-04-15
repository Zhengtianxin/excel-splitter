#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 拆分工具 v1.1
按指定列拆分xlsx文件，保留原表所有格式。
支持打包为 Windows 单文件 exe。
"""

import os
import sys
import glob
import copy
import traceback
from collections import defaultdict

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("=" * 60)
    print("  错误：缺少 openpyxl 库！")
    print("  请执行以下命令安装：")
    print("    pip install openpyxl")
    print("=" * 60)
    wait_exit()
    sys.exit(1)


# ==================== 退出等待 ====================

def wait_exit():
    """按任意键退出（Windows用msvcrt，其他系统用input兜底）"""
    print()
    try:
        import msvcrt
        print("按任意键退出...", end='', flush=True)
        msvcrt.getch()
    except ImportError:
        input("按回车键退出...")
    print()


# ==================== 样式复制工具函数 ====================

def copy_cell_style(src_cell, dst_cell):
    """完整复制单元格样式（字体/填充/边框/对齐/保护/数字格式）"""
    try:
        if src_cell.font:
            dst_cell.font = copy.copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = copy.copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy.copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = copy.copy(src_cell.alignment)
        if src_cell.protection:
            dst_cell.protection = copy.copy(src_cell.protection)
        if src_cell.number_format:
            dst_cell.number_format = src_cell.number_format
    except Exception:
        pass  # 样式复制失败不影响数据


def copy_cell(src_cell, dst_cell):
    """完整复制单元格（值 + 样式 + 超链接 + 批注）"""
    dst_cell.value = src_cell.value
    copy_cell_style(src_cell, dst_cell)
    try:
        if src_cell.hyperlink:
            dst_cell.hyperlink = copy.copy(src_cell.hyperlink)
    except Exception:
        pass
    try:
        if src_cell.comment:
            dst_cell.comment = copy.copy(src_cell.comment)
    except Exception:
        pass


def copy_sheet_properties(src_ws, dst_ws):
    """复制工作表级别属性（列宽/冻结窗格/页面设置等）"""
    # 列宽
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        try:
            if col_dim.font:
                dst_ws.column_dimensions[col_letter].font = copy.copy(col_dim.font)
            if col_dim.fill:
                dst_ws.column_dimensions[col_letter].fill = copy.copy(col_dim.fill)
            if col_dim.border:
                dst_ws.column_dimensions[col_letter].border = copy.copy(col_dim.border)
            if col_dim.alignment:
                dst_ws.column_dimensions[col_letter].alignment = copy.copy(col_dim.alignment)
        except Exception:
            pass

    # 冻结窗格
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass

    # 页面设置
    try:
        dst_ws.page_setup.orientation = src_ws.page_setup.orientation
        dst_ws.page_setup.paperSize = src_ws.page_setup.paperSize
        dst_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight
        dst_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
    except Exception:
        pass

    # 打印标题行/列
    try:
        if src_ws.print_title_rows:
            dst_ws.print_title_rows = src_ws.print_title_rows
        if src_ws.print_title_cols:
            dst_ws.print_title_cols = src_ws.print_title_cols
    except Exception:
        pass

    # 自动筛选
    try:
        if src_ws.auto_filter and src_ws.auto_filter.ref:
            dst_ws.auto_filter.ref = src_ws.auto_filter.ref
    except Exception:
        pass

    # Tab颜色
    try:
        if src_ws.sheet_properties.tabColor:
            dst_ws.sheet_properties.tabColor = copy.copy(src_ws.sheet_properties.tabColor)
    except Exception:
        pass


# ==================== 显示工具函数 ====================

def truncate(text, max_len=12):
    """截断过长文本用于显示"""
    if text is None:
        return "(空)"
    text = str(text).replace('\n', ' ').replace('\r', '')
    if len(text) > max_len:
        return text[:max_len] + "…"
    return text


def format_row_preview(ws, row_idx, max_col):
    """格式化一行的预览内容，用 | 分隔各列"""
    parts = []
    # 最多显示前10列，避免太宽
    display_cols = min(max_col, 10)
    for col_idx in range(1, display_cols + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        parts.append(truncate(val))
    preview = " | ".join(parts)
    if max_col > 10:
        preview += f" | …(共{max_col}列)"
    return preview


def safe_filename(name):
    """清理文件名中的非法字符"""
    if not name:
        return "未分类"
    # Windows 文件名非法字符
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, '_')
    # 去除首尾空格和点（Windows不允许文件名以点结尾）
    name = name.strip().rstrip('.')
    if not name:
        return "未分类"
    # 限制长度（Windows路径最长260，文件名保守限制100）
    if len(name) > 100:
        name = name[:100]
    return name


# ==================== 嵌套检测 ====================

def detect_nested_values(group_names):
    """
    检测分组名称中是否存在嵌套/包含关系。
    例如："通州" 包含于 "北京通州"，则提示用户。
    返回: [(短名, 长名), ...] 有包含关系的配对列表
    """
    names = sorted(group_names, key=len)
    nested_pairs = []
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            short, long = names[i], names[j]
            if short == long:
                continue
            if short in long:
                nested_pairs.append((short, long))
    return nested_pairs


# ==================== 主逻辑 ====================

def find_xlsx_file(work_dir):
    """在指定目录下查找xlsx文件"""
    xlsx_files = glob.glob(os.path.join(work_dir, '*.xlsx'))
    if not xlsx_files:
        return None, "当前目录下没有找到任何 .xlsx 文件！"

    if len(xlsx_files) == 1:
        return xlsx_files[0], None

    # 多个文件，让用户选择
    print("\n当前目录下有多个 .xlsx 文件，请选择要拆分的文件：")
    print("-" * 50)
    for i, f in enumerate(xlsx_files):
        size = os.path.getsize(f)
        if size > 1024 * 1024:
            size_str = f"{size / 1024 / 1024:.1f} MB"
        else:
            size_str = f"{size / 1024:.1f} KB"
        print(f"  {i + 1}. {os.path.basename(f)}  ({size_str})")
    print("-" * 50)

    while True:
        choice = input(f"请输入文件编号 (1-{len(xlsx_files)}): ").strip()
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(xlsx_files):
                return xlsx_files[idx], None
            else:
                print(f"  ✗ 请输入 1 到 {len(xlsx_files)} 之间的数字")
        except ValueError:
            print(f"  ✗ 请输入数字，不要输入其他字符")


def ask_header_row(ws, max_col):
    """交互式询问标题行"""
    max_row = ws.max_row
    preview_rows = min(max_row, 5)  # 预览前5行

    print("\n" + "=" * 60)
    print("  第一步：请选择【标题行】是哪一行")
    print("  (数据行将从标题行的下一行自动开始)")
    print("=" * 60)
    print()

    # 展示前几行的内容
    for r in range(1, preview_rows + 1):
        preview = format_row_preview(ws, r, max_col)
        print(f"  {r}. 第{r}行：{preview}")

    if max_row > preview_rows:
        print(f"\n  (共 {max_row} 行，以上仅展示前 {preview_rows} 行)")

    print(f"\n  0. 其他 → 手动输入行号")
    print()

    while True:
        choice = input("请选择标题行编号: ").strip()
        try:
            num = int(choice)
            if num == 0:
                custom = input(f"请输入标题行的行号 (1-{max_row}): ").strip()
                try:
                    custom_num = int(custom)
                    if 1 <= custom_num <= max_row:
                        return custom_num
                    else:
                        print(f"  ✗ 行号超出范围，表格共 {max_row} 行，请输入 1 到 {max_row}")
                except ValueError:
                    print("  ✗ 请输入数字")
            elif 1 <= num <= preview_rows:
                return num
            else:
                print(f"  ✗ 请输入 0 到 {preview_rows} 之间的数字")
        except ValueError:
            print("  ✗ 请输入数字，不要输入其他字符")


def ask_split_column(ws, header_row, max_col):
    """交互式询问按哪列拆分"""
    print("\n" + "=" * 60)
    print(f"  第二步：请选择按哪一列拆分")
    print(f"  (以下是第 {header_row} 行标题行的各列内容)")
    print("=" * 60)
    print()

    # 展示标题行每一列
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        val = ws.cell(row=header_row, column=col_idx).value
        display = str(val) if val is not None else "(空)"
        # 截断过长的
        if len(display) > 30:
            display = display[:30] + "…"
        print(f"  {col_idx}. {col_letter}列：{display}")

    print()

    while True:
        choice = input(f"请选择要拆分的列号 (1-{max_col}): ").strip()
        try:
            num = int(choice)
            if 1 <= num <= max_col:
                col_letter = get_column_letter(num)
                val = ws.cell(row=header_row, column=num).value
                display = str(val) if val is not None else "(空)"
                print(f"\n  ✓ 已选择：{col_letter}列 -「{display}」")
                return num
            else:
                print(f"  ✗ 请输入 1 到 {max_col} 之间的数字，共 {max_col} 列")
        except ValueError:
            print("  ✗ 请输入数字，不要输入其他字符")


def do_split(src_file, ws, header_row, split_col, max_row, max_col):
    """执行拆分"""
    work_dir = os.path.dirname(os.path.abspath(src_file))
    data_start_row = header_row + 1

    if data_start_row > max_row:
        print(f"\n  ✗ 错误：标题行是第 {header_row} 行，但表格只有 {max_row} 行，没有数据行可拆分！")
        return False

    col_letter = get_column_letter(split_col)
    col_name = ws.cell(row=header_row, column=split_col).value or "(空)"

    print("\n" + "=" * 60)
    print("  开始拆分")
    print("=" * 60)
    print(f"  源文件：{os.path.basename(src_file)}")
    print(f"  标题行：第 {header_row} 行")
    print(f"  数据行：第 {data_start_row} 行 ~ 第 {max_row} 行（共 {max_row - data_start_row + 1} 行数据）")
    print(f"  拆分列：{col_letter}列 -「{col_name}」")
    print("-" * 60)

    # ========== 按指定列分组 ==========
    groups = defaultdict(list)
    empty_count = 0
    for row_idx in range(data_start_row, max_row + 1):
        cell_value = ws.cell(row=row_idx, column=split_col).value
        if cell_value is None or str(cell_value).strip() == '':
            key = "未分类"
            empty_count += 1
        else:
            key = str(cell_value).strip()
        groups[key].append(row_idx)

    if empty_count > 0:
        print(f"\n  ⚠ 注意：有 {empty_count} 行数据的{col_letter}列为空，将归入「未分类.xlsx」")

    print(f"\n  共将拆分为 {len(groups)} 个文件：")
    for name, rows in groups.items():
        print(f"    • {name} → {len(rows)} 行数据")

    # ========== 嵌套检测 ==========
    real_names = [n for n in groups.keys() if n != "未分类"]
    nested_pairs = detect_nested_values(real_names)
    if nested_pairs:
        print()
        print("  " + "!" * 56)
        print("  ⚠ 警告：以下分组名称之间存在【包含/嵌套】关系：")
        print("  （它们会被拆分成不同的文件，请确认这是您期望的结果）")
        print("  " + "-" * 56)
        for short, long in nested_pairs:
            s_count = len(groups[short])
            l_count = len(groups[long])
            print(f"    「{short}」({s_count}行) ⊂ 「{long}」({l_count}行)")
        print("  " + "-" * 56)
        print(f"  如果「{nested_pairs[0][0]}」和「{nested_pairs[0][1]}」本应是同一类，")
        print(f"  请先在原表中统一{col_letter}列的值，再重新运行此工具。")
        print("  " + "!" * 56)
        print()
        confirm = input("  是否继续拆分？(Y/n): ").strip().lower()
        if confirm == 'n':
            print("\n  已取消拆分。")
            return False

    # ========== 收集合并单元格信息 ==========
    # 标题区域的合并（第1行到标题行）
    header_merged = []
    for mg in ws.merged_cells.ranges:
        if mg.min_row >= 1 and mg.max_row <= header_row:
            header_merged.append(mg)

    # 数据行中的同行合并
    data_row_merges = defaultdict(list)
    for mg in ws.merged_cells.ranges:
        if mg.min_row >= data_start_row and mg.min_row == mg.max_row:
            data_row_merges[mg.min_row].append((mg.min_col, mg.max_col))

    # 收集行高
    row_heights = {}
    for row_idx in range(1, max_row + 1):
        rd = ws.row_dimensions.get(row_idx)
        if rd and rd.height:
            row_heights[row_idx] = rd.height

    # ========== 逐组生成文件 ==========
    print("\n  正在生成文件...")
    success_count = 0
    fail_count = 0
    generated_files = []

    for group_name, src_rows in groups.items():
        fname = safe_filename(group_name)
        dst_file = os.path.join(work_dir, f"{fname}.xlsx")

        try:
            dst_wb = Workbook()
            dst_ws = dst_wb.active
            dst_ws.title = ws.title

            # 复制sheet属性
            copy_sheet_properties(ws, dst_ws)

            # 复制标题区域（第1行到header_row）
            for h_row in range(1, header_row + 1):
                for col_idx in range(1, max_col + 1):
                    src_cell = ws.cell(row=h_row, column=col_idx)
                    dst_cell = dst_ws.cell(row=h_row, column=col_idx)
                    copy_cell(src_cell, dst_cell)
                if h_row in row_heights:
                    dst_ws.row_dimensions[h_row].height = row_heights[h_row]

            # 标题区域合并单元格
            for mg in header_merged:
                try:
                    dst_ws.merge_cells(
                        start_row=mg.min_row, start_column=mg.min_col,
                        end_row=mg.max_row, end_column=mg.max_col
                    )
                except Exception:
                    pass

            # 复制数据行
            for offset, src_row_idx in enumerate(src_rows):
                dst_row_idx = header_row + 1 + offset

                for col_idx in range(1, max_col + 1):
                    src_cell = ws.cell(row=src_row_idx, column=col_idx)
                    dst_cell = dst_ws.cell(row=dst_row_idx, column=col_idx)
                    copy_cell(src_cell, dst_cell)

                if src_row_idx in row_heights:
                    dst_ws.row_dimensions[dst_row_idx].height = row_heights[src_row_idx]

                if src_row_idx in data_row_merges:
                    for min_c, max_c in data_row_merges[src_row_idx]:
                        try:
                            dst_ws.merge_cells(
                                start_row=dst_row_idx, start_column=min_c,
                                end_row=dst_row_idx, end_column=max_c
                            )
                        except Exception:
                            pass

            # 再次确保列宽精确
            for col_idx in range(1, max_col + 1):
                cl = get_column_letter(col_idx)
                src_dim = ws.column_dimensions.get(cl)
                if src_dim:
                    dst_ws.column_dimensions[cl].width = src_dim.width

            # 数据验证
            try:
                for dv in ws.data_validations.dataValidation:
                    dst_ws.add_data_validation(copy.copy(dv))
            except Exception:
                pass

            dst_wb.save(dst_file)
            print(f"    ✓ {fname}.xlsx（{len(src_rows)} 行）")
            success_count += 1
            generated_files.append(f"{fname}.xlsx")

        except PermissionError:
            print(f"    ✗ {fname}.xlsx → 失败！文件被占用，请关闭该文件后重试")
            print(f"      详情：Windows 下如果该文件正在被 Excel 打开，会导致无法写入。")
            print(f"      解决：关闭 Excel 中的 {fname}.xlsx，然后重新运行此工具。")
            fail_count += 1
        except OSError as e:
            print(f"    ✗ {fname}.xlsx → 失败！系统错误")
            print(f"      错误代码：{e.errno}")
            print(f"      错误详情：{e}")
            if 'name too long' in str(e).lower() or e.errno == 36:
                print(f"      原因分析：文件名过长。原始值「{group_name}」超出系统限制。")
            elif 'no space' in str(e).lower() or e.errno == 28:
                print(f"      原因分析：磁盘空间不足，请清理磁盘后重试。")
            else:
                print(f"      原因分析：可能是磁盘权限问题或路径问题。")
            fail_count += 1
        except Exception as e:
            print(f"    ✗ {fname}.xlsx → 失败！")
            print(f"      错误类型：{type(e).__name__}")
            print(f"      错误详情：{e}")
            fail_count += 1

    # ========== 结果汇总 ==========
    print()
    print("=" * 60)
    if fail_count == 0:
        print(f"  ✓ 拆分完成！成功生成 {success_count} 个文件")
    else:
        print(f"  ⚠ 拆分完成：成功 {success_count} 个，失败 {fail_count} 个")
    print(f"  文件保存在：{work_dir}")
    print("=" * 60)

    # 再次提醒嵌套
    if nested_pairs:
        print()
        print("  ⚠ 再次提醒：以下名称存在包含关系，已拆分为不同文件：")
        for short, long in nested_pairs:
            print(f"    「{short}」 和 「{long}」 → 分别生成了独立文件")
        print("  如果不符合预期，请在原表中统一名称后重新拆分。")

    return fail_count == 0


def main():
    print()
    print("+" + "=" * 46 + "+")
    print("|        Excel 拆分工具  v1.1                 |")
    print("|  按指定列拆分xlsx，保留原表所有格式         |")
    print("+" + "=" * 46 + "+")
    print()

    # ========== 确定工作目录 ==========
    # 如果是 PyInstaller 打包的 exe，工作目录为 exe 所在目录
    if getattr(sys, 'frozen', False):
        work_dir = os.path.dirname(sys.executable)
    else:
        work_dir = os.path.dirname(os.path.abspath(__file__))

    print(f"工作目录：{work_dir}")

    # ========== 查找xlsx文件 ==========
    src_file, err = find_xlsx_file(work_dir)
    if err:
        print(f"\n  ✗ {err}")
        print(f"  请将此工具放到 .xlsx 文件所在的文件夹中，再运行。")
        wait_exit()
        sys.exit(1)

    print(f"已找到文件：{os.path.basename(src_file)}")

    # ========== 打开文件 ==========
    print(f"正在读取，请稍候...")
    try:
        wb = load_workbook(src_file)
    except PermissionError:
        print(f"\n  ✗ 无法打开文件！文件可能正被 Excel 占用。")
        print(f"  请先关闭 Excel 中打开的该文件，再重新运行此工具。")
        wait_exit()
        sys.exit(1)
    except Exception as e:
        print(f"\n  ✗ 读取文件失败！")
        print(f"  文件路径：{src_file}")
        print(f"  错误类型：{type(e).__name__}")
        print(f"  错误详情：{e}")
        print(f"\n  可能原因：")
        print(f"    1. 文件不是有效的 .xlsx 格式（可能是 .xls 旧格式，需另存为 .xlsx）")
        print(f"    2. 文件已损坏（尝试用 Excel 打开并重新保存）")
        print(f"    3. 文件有密码保护（请先去除密码）")
        print(f"\n  完整错误堆栈：")
        print("-" * 60)
        traceback.print_exc()
        print("-" * 60)
        wait_exit()
        sys.exit(1)

    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row is None or max_row < 1:
        print(f"\n  ✗ 表格为空，没有任何数据！")
        wait_exit()
        sys.exit(1)

    if max_col is None or max_col < 1:
        print(f"\n  ✗ 表格没有任何列！")
        wait_exit()
        sys.exit(1)

    print(f"读取成功！表格共 {max_row} 行，{max_col} 列，Sheet名称：「{ws.title}」\n")

    if max_row < 2:
        print(f"\n  ✗ 表格只有 {max_row} 行，至少需要1行标题 + 1行数据才能拆分！")
        wait_exit()
        sys.exit(1)

    # ========== 交互问答 ==========
    header_row = ask_header_row(ws, max_col)
    print(f"\n  ✓ 标题行：第 {header_row} 行")
    print(f"  ✓ 数据行：从第 {header_row + 1} 行开始")

    split_col = ask_split_column(ws, header_row, max_col)

    # ========== 执行拆分 ==========
    try:
        do_split(src_file, ws, header_row, split_col, max_row, max_col)
    except Exception as e:
        print(f"\n  ✗ 拆分过程中发生未预期的错误！")
        print(f"  错误类型：{type(e).__name__}")
        print(f"  错误详情：{e}")
        print(f"\n  完整错误堆栈（可截图发给开发者排查）：")
        print("-" * 60)
        traceback.print_exc()
        print("-" * 60)

    wait_exit()


if __name__ == '__main__':
    main()
